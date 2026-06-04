"""
utils/helpers.py
────────────────
Funções utilitárias:
  - Validação de CPF
  - Geração de QR Code
  - Formatação de dados
  - Exportação CSV / Excel
"""

import os
import re
import io
from datetime import datetime
from typing import Optional

# ═══════════════════════════════════════════════════════════════════
# URL BASE
# ═══════════════════════════════════════════════════════════════════

def get_app_url() -> str:
    """Retorna a URL base da aplicação."""
    # 1. Variável de ambiente
    url = os.getenv("APP_URL", "").strip().rstrip("/")
    if url:
        return url
    # 2. Streamlit Secrets
    try:
        import streamlit as st
        url = st.secrets.get("APP_URL", "").strip().rstrip("/")
        if url:
            return url
    except Exception:
        pass
    # 3. Fallback local
    return "http://localhost:8501"


def link_presenca(tid: str) -> str:
    return f"{get_app_url()}/?page=presenca&id={tid}"


# ═══════════════════════════════════════════════════════════════════
# VALIDAÇÃO DE CPF
# ═══════════════════════════════════════════════════════════════════

def validar_cpf(cpf: str) -> bool:
    """Valida CPF (algoritmo completo com dígitos verificadores)."""
    digits = re.sub(r"\D", "", cpf)
    if len(digits) != 11 or len(set(digits)) == 1:
        return False
    # Primeiro dígito
    s = sum(int(digits[i]) * (10 - i) for i in range(9))
    r = (s * 10) % 11
    if r in (10, 11):
        r = 0
    if r != int(digits[9]):
        return False
    # Segundo dígito
    s = sum(int(digits[i]) * (11 - i) for i in range(10))
    r = (s * 10) % 11
    if r in (10, 11):
        r = 0
    return r == int(digits[10])


def formatar_cpf(cpf: str) -> str:
    d = re.sub(r"\D", "", cpf)
    if len(d) != 11:
        return cpf
    return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"


def mascarar_cpf(cpf: str) -> str:
    """Exibe CPF com dígitos centrais mascarados (para relatórios públicos)."""
    d = re.sub(r"\D", "", cpf)
    if len(d) != 11:
        return cpf
    return f"{d[:3]}.***.***-{d[9:]}"


def validar_email(email: str) -> bool:
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email.strip()))


def nome_completo(nome: str) -> bool:
    """Verifica se há pelo menos 2 palavras."""
    return len([w for w in nome.strip().split() if w]) >= 2


# ═══════════════════════════════════════════════════════════════════
# QR CODE
# ═══════════════════════════════════════════════════════════════════

def gerar_qr_bytes(url: str, tamanho: int = 300) -> bytes:
    """Gera QR Code como bytes PNG."""
    import qrcode
    from qrcode.image.pure import PyPNGImage

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)

    try:
        from PIL import Image as PilImage
        img = qr.make_image(fill_color="#0f0e0c", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        img = qr.make_image(image_factory=PyPNGImage)
        buf = io.BytesIO()
        img.save(buf)
        return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════
# FORMATAÇÃO
# ═══════════════════════════════════════════════════════════════════

def fmt_data(val) -> str:
    if not val:
        return "—"
    if hasattr(val, "strftime"):
        return val.strftime("%d/%m/%Y")
    try:
        return datetime.fromisoformat(str(val)).strftime("%d/%m/%Y")
    except Exception:
        return str(val)


def fmt_hora(val) -> str:
    if not val:
        return ""
    return str(val)[:5]


def fmt_datetime(val) -> str:
    if not val:
        return "—"
    try:
        return datetime.fromisoformat(str(val)).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(val)


def fmt_carga(val) -> str:
    if not val:
        return "—"
    return f"{float(val):.0f}h"


# ═══════════════════════════════════════════════════════════════════
# EXPORTAÇÃO
# ═══════════════════════════════════════════════════════════════════

def exportar_csv(treinamento: dict, presencas: list[dict]) -> bytes:
    """Gera CSV com metadados do treinamento + lista de presenças."""
    import csv

    buf = io.StringIO()
    w = csv.writer(buf)

    # Cabeçalho auditável
    w.writerow(["# RELATÓRIO DE PRESENÇA — PresençaCerta"])
    w.writerow(["# Treinamento", treinamento.get("nome", "")])
    w.writerow(["# Data", fmt_data(treinamento.get("data")) + " " + fmt_hora(treinamento.get("hora"))])
    w.writerow(["# Local", treinamento.get("local") or "—"])
    w.writerow(["# Setor Organizador", treinamento.get("setor") or "—"])
    w.writerow(["# Instrutor", treinamento.get("instrutor") or "—"])
    w.writerow(["# Carga Horária", fmt_carga(treinamento.get("carga_horas"))])
    w.writerow(["# Total de Presenças", len(presencas)])
    w.writerow(["# Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")])
    w.writerow([])

    # Tabela
    w.writerow(["#", "Nome Completo", "CPF", "E-mail Corporativo", "Cargo/Função", "Matrícula", "Data Registro", "Hora Registro"])
    for i, p in enumerate(presencas, 1):
        dt = fmt_datetime(p.get("registrado_em"))
        parts = dt.split(" ") if dt != "—" else ["—", ""]
        w.writerow([
            i,
            p.get("nome", ""),
            formatar_cpf(p.get("cpf", "")),
            p.get("email", ""),
            p.get("cargo") or "—",
            p.get("matricula") or "—",
            parts[0],
            parts[1] if len(parts) > 1 else "",
        ])

    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def exportar_excel(treinamento: dict, presencas: list[dict]) -> bytes:
    """Gera planilha Excel formatada."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presenças"

    INK    = "0F0E0C"
    ACCENT = "C8440A"
    PAPER  = "F5F0E8"
    CREAM  = "EDE8DC"

    # ── Cabeçalho do relatório ──
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"RELATÓRIO DE PRESENÇA — {treinamento.get('nome', '').upper()}"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=INK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    meta = [
        ("Data", fmt_data(treinamento.get("data")) + " " + fmt_hora(treinamento.get("hora"))),
        ("Local", treinamento.get("local") or "—"),
        ("Setor Organizador", treinamento.get("setor") or "—"),
        ("Instrutor", treinamento.get("instrutor") or "—"),
        ("Carga Horária", fmt_carga(treinamento.get("carga_horas"))),
        ("Total de Presenças", str(len(presencas))),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for row_idx, (label, val) in enumerate(meta, start=2):
        ws.merge_cells(f"A{row_idx}:C{row_idx}")
        lc = ws.cell(row=row_idx, column=1, value=label)
        lc.font = Font(name="Calibri", bold=True, color=INK)
        lc.fill = PatternFill("solid", fgColor=CREAM)

        ws.merge_cells(f"D{row_idx}:H{row_idx}")
        vc = ws.cell(row=row_idx, column=4, value=val)
        vc.font = Font(name="Calibri", color=INK)

    # ── Linha de separação ──
    sep_row = len(meta) + 3
    ws.merge_cells(f"A{sep_row}:H{sep_row}")
    sep = ws[f"A{sep_row}"]
    sep.fill = PatternFill("solid", fgColor=ACCENT)
    ws.row_dimensions[sep_row].height = 4

    # ── Cabeçalho da tabela ──
    header_row = sep_row + 1
    headers = ["#", "Nome Completo", "CPF", "E-mail Corporativo", "Cargo/Função", "Matrícula", "Data Registro", "Hora Registro"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="FFFFFF"))
    ws.row_dimensions[header_row].height = 20

    # ── Dados ──
    thin = Side(style="thin", color="D4CEC2")
    border = Border(bottom=thin)

    for i, p in enumerate(presencas):
        r = header_row + 1 + i
        dt = fmt_datetime(p.get("registrado_em"))
        parts = dt.split(" ") if dt != "—" else ["—", ""]
        row_data = [
            i + 1,
            p.get("nome", ""),
            formatar_cpf(p.get("cpf", "")),
            p.get("email", ""),
            p.get("cargo") or "—",
            p.get("matricula") or "—",
            parts[0],
            parts[1] if len(parts) > 1 else "",
        ]
        fill = PatternFill("solid", fgColor=PAPER if i % 2 == 0 else "FFFFFF")
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Calibri", size=10)
            c.fill = fill
            c.border = border
            if col == 1:
                c.alignment = Alignment(horizontal="center")

    # ── Larguras das colunas ──
    col_widths = [5, 35, 18, 35, 25, 15, 16, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ── Rodapé ──
    footer_row = header_row + len(presencas) + 2
    ws.merge_cells(f"A{footer_row}:H{footer_row}")
    fc = ws[f"A{footer_row}"]
    fc.value = f"Documento gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} — PresençaCerta"
    fc.font = Font(name="Calibri", size=8, italic=True, color="7A7268")
    fc.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
